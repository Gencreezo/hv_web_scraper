from Adresse import Adresse
from Customer import Customer
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl import workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import time
import xlrd


plz = "30823"
city = "Garbsen"

loc = ("Adressen_Garbsen.xlsx")

wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)

addresses = []
customers = []


customerWB = load_workbook('Kunden_Liste.xlsx')
custSheets = customerWB.sheetnames
custSheet = customerWB[custSheets[0]]

adressenWB = load_workbook('Adressen_Liste.xlsx')
adrSheets = adressenWB.sheetnames
adrSheet = adressenWB[adrSheets[0]]

# lastIndex = adrSheet.cell(row=1, column=1).value  # 572

lastIndex = 1139

driver = webdriver.Chrome()


def set_index(index):
    lastIndex = index


def get_index():
    return lastIndex


def write_customer_to_xlsx(customerData, adresse):
    soup = BeautifulSoup(driver.page_source, 'html.parser')

    kunde_seit = customerData[2]
    kunden_nr = customerData[3]
    ka = customerData[4]
    kd = customerData[5]
    ki_p = customerData[6]
    status = customerData[7]

    anrede = soup.find('option', selected=True).get_text()

    nachname = soup.find('input', {'id': 'lastName'}).get('value')

    vorname = soup.find('input', {'id': 'firstName'}).get('value')

    geburtstag = soup.find('input', {'id': 'birthday'}).get('value')

    # Telefon privat
    telPrivateAreaCode = soup.find(
        'input', {'id': 'telPrivateAreaCode'}).get('value')

    telPrivateExchange = soup.find(
        'input', {'id': 'telPrivateExchange'}).get('value')

    if telPrivateAreaCode or telPrivateExchange:
        telefonPrivat = telPrivateAreaCode + " / " + telPrivateExchange
    else:
        telefonPrivat = ""

    # Telefon dienstlich
    telWorkAreaCode = soup.find(
        'input', {'id': 'telWorkAreaCode'}).get('value')

    telWorkExchange = soup.find(
        'input', {'id': 'telWorkExchange'}).get('value')

    if telWorkAreaCode or telWorkExchange:
        telefonWork = telWorkAreaCode + " / " + telWorkExchange
    else:
        telefonWork = ""

    # Telefon mobil
    telMobilAreaCode = soup.find(
        'input', {'id': 'telMobilAreaCode'}).get('value')

    telMobilExchange = soup.find(
        'input', {'id': 'telMobilExchange'}).get('value')

    if telMobilAreaCode or telMobilExchange:
        telefonMobil = telMobilAreaCode + " / " + telMobilExchange
    else:
        telefonMobil = ""

    # email
    email = soup.find('input', {'id': 'email'}).get('value')

    # Verfügbarkeit
    a = soup.find('div', {'id': 'ampelA'}).get('class')
    if a[0] == "marketable":
        a = 'GRÜN'
    else:
        a = 'ROT'

    d = soup.find('div', {'id': 'ampelD'}).get('class')
    if d[0] == "marketable":
        d = 'GRÜN'
    else:
        d = 'ROT'

    ip = soup.find('div', {'id': 'ampelIP'}).get('class')
    if ip[0] == "marketable":
        ip = 'GRÜN'
    else:
        ip = 'ROT'

    ud = soup.find('div', {'id': 'ampelUD'}).get('class')
    if ud[0] == "marketable":
        ud = 'GRÜN'
    else:
        ud = 'ROT'

    customer = Customer(adresse, anrede, nachname, vorname, geburtstag, telefonPrivat, telefonWork,
                        telefonMobil, email, kunde_seit, kunden_nr, ka, kd, ki_p, status, a, d, ip, ud)
    customers.append(customer)


def login():
    driver.get("https://secure.kabeldeutschland.de/hv")
    driver.maximize_window()
    elem = driver.find_element_by_name("IDToken1")
    elem.clear()
    elem.send_keys(username)
    elem = driver.find_element_by_name("IDToken2")
    elem.send_keys(password)
    elem.send_keys(Keys.TAB)
    elem.send_keys(Keys.ENTER)
    time.sleep(90)
    print('Warnung: Noch 60 Sekunden')
    time.sleep(50)
    print('Warnung: Noch 10 Sekunden')
    time.sleep(10)


def read_addresses():
    for row in range(sheet.nrows):
        street = sheet.cell_value(row, 0)
        nummer = sheet.cell_value(row, 1)
        zusatz = sheet.cell_value(row, 2)
        adr = Adresse(plz, city, street, nummer, zusatz)
        addresses.append(adr)


def loop():
    for address in addresses[get_index():]:
        try:
            adresse = address.plz + '; ' + address.ort + '; ' + \
                address.strasse + '; ' + address.nummer + '; ' + address.zusatz

            driver.get("***") # link censored because of security reasons

            elem = driver.find_elements_by_id('PIN')
            if elem:
                print(adresse)
                print(addresses.index(address))
                break

            elem = driver.find_element_by_id("plzort[plz]")
            elem.clear()
            elem.send_keys(plz)
            elem = driver.find_element_by_id("plzort[ortsname]")
            elem.clear()
            elem.send_keys(city)

            elem = driver.find_element_by_id('strasse')
            elem.clear()
            elem.send_keys(address.strasse)

            elem = driver.find_element_by_id('hausNr')
            elem.clear()
            elem.send_keys(address.nummer)

            if address.zusatz:
                elem = driver.find_element_by_id('hausNrZusatz')
                elem.clear()
                elem.send_keys(address.zusatz)
            else:
                elem = driver.find_element_by_id('hausNrZusatz')
                elem.clear()

            elem = driver.find_element_by_link_text(
                "Adresse ergänzen/prüfen").click()

            elem = driver.find_elements_by_name('adrList')
            if elem:
                print('Found "adrList"!')
                continue

            table = driver.find_elements_by_class_name('textMargin')
            if table:
                rows = table[0].find_elements_by_tag_name('tr')
                rows.pop(0)
                for i in range(len(rows)+1):
                    try:
                        elem = driver.find_elements_by_tag_name('tr')[i]
                        rowText = elem.text
                        arr = rowText.split('  ')
                        elem.click()
                        elem = driver.find_elements_by_class_name('fehlerinfo')
                        if elem:
                            driver.get(
                                "***")
                            elem = driver.find_element_by_link_text(
                                "Adresse ergänzen/prüfen").click()
                            continue
                        write_customer_to_xlsx(arr, adresse)
                        elem = driver.find_element_by_link_text(
                            "Zur Kundensuche").click()
                        elem = driver.find_element_by_link_text(
                            "Adresse ergänzen/prüfen").click()
                        rows = table[0].find_elements_by_tag_name('tr')
                    except:
                        continue
            else:
                addresses.pop(address)
        except:
            elem = driver.find_elements_by_name('serviceAddressList')
            print(elem)
            if not elem:
                print('Letzter Index: ', addresses.index(address))
                set_index(addresses.index(address))
                break
            continue


def write_customers():
    wRow = custSheet.max_row + 1
    print('Customers starting at Row: ', wRow)

    for customer in customers:
        custSheet.cell(row=wRow, column=1).value = customer.adresse
        custSheet.cell(row=wRow, column=2).value = customer.anrede
        custSheet.cell(row=wRow, column=3).value = customer.nachname
        custSheet.cell(row=wRow, column=4).value = customer.vorname
        custSheet.cell(row=wRow, column=5).value = customer.geburtstag
        custSheet.cell(row=wRow, column=6).value = customer.telefonPrivat
        custSheet.cell(row=wRow, column=7).value = customer.telefonWork
        custSheet.cell(row=wRow, column=8).value = customer.telefonMobil
        custSheet.cell(row=wRow, column=9).value = customer.email
        custSheet.cell(row=wRow, column=11).value = customer.kunde_seit
        custSheet.cell(row=wRow, column=12).value = customer.kunden_nr
        custSheet.cell(row=wRow, column=13).value = customer.ka
        custSheet.cell(row=wRow, column=14).value = customer.kd
        custSheet.cell(row=wRow, column=15).value = customer.ki_p
        custSheet.cell(row=wRow, column=16).value = customer.status
        custSheet.cell(row=wRow, column=18).value = customer.a
        custSheet.cell(row=wRow, column=19).value = customer.d
        custSheet.cell(row=wRow, column=20).value = customer.ip
        custSheet.cell(row=wRow, column=21).value = customer.ud
        wRow += 1


def updateAddressList():
    adrSheet.cell(row=1, column=1).value = get_index()
    row = 2
    for address in addresses:
        adrSheet.cell(row=row, column=1).value = address.strasse
        adrSheet.cell(row=row, column=2).value = address.nummer
        adrSheet.cell(row=row, column=3).value = address.zusatz
        row += 1


def main():
    print('Starting at Index: ', lastIndex)
    read_addresses()
    login()
    loop()

    updateAddressList()
    adressenWB.save('Adressen_Liste.xlsx')
    adressenWB.close()

    write_customers()
    customerWB.save('Kunden_Liste.xlsx')
    customerWB.close()

    driver.close()


main()
