from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Kunden_Garbsen"

wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Adressen_Garbsen"


def init_cust_table():
    ws.cell(row=1, column=1).value = 'Adresse'
    ws.cell(row=1, column=2).value = 'Anrede'
    ws.cell(row=1, column=3).value = 'Nachname'
    ws.cell(row=1, column=4).value = 'Vorname'
    ws.cell(row=1, column=5).value = 'Geburtstag'
    ws.cell(row=1, column=6).value = 'Telefon (Privat)'
    ws.cell(row=1, column=7).value = 'Telefon (Work)'
    ws.cell(row=1, column=8).value = 'Telefon (Mobil)'
    ws.cell(row=1, column=9).value = 'E-Mail'
    ws.cell(row=1, column=10).value = 'IBAN'
    ws.cell(row=1, column=11).value = 'Kunde seit'
    ws.cell(row=1, column=12).value = 'Kundennr.'
    ws.cell(row=1, column=13).value = 'KA'
    ws.cell(row=1, column=14).value = 'KD'
    ws.cell(row=1, column=15).value = 'KI/P'
    ws.cell(row=1, column=16).value = 'Status'
    ws.cell(row=1, column=17).value = 'Verfügbarkeit'
    ws.cell(row=1, column=18).value = 'A'
    ws.cell(row=1, column=19).value = 'D'
    ws.cell(row=1, column=20).value = 'I/P'
    ws.cell(row=1, column=21).value = 'UD'


def init_adr_table():
    ws.cell(row=1, column=1).value = 'Straße'
    ws.cell(row=1, column=2).value = 'Nummer'
    ws.cell(row=1, column=3).value = 'Zusatz'


init_cust_table()
wb.save(filename='Kunden_Liste.xlsx')
wb.close

init_adr_table()
wb2.save(filename='Adressen_Liste.xlsx')
wb2.close
